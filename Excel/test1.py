from socket import if_indextoname
from turtle import pos
import openpyxl
import string
from pyrsistent import v
from sympy import li, true


workbook = openpyxl.load_workbook(r'C:\Users\Rajo\Develop\Excel\primer ponuda.xlsx')
sheets = workbook.sheetnames 
print(sheets)
lst = ['opis radova']


ii=0
pos_list = []
for x in sheets:
    #print(x)
    workbook.active = ii
    #print(workbook.active,ii)
    sheet1 = workbook[workbook.active.title]
    rows = sheet1.max_row
    columns = sheet1.max_column
    ii +=1
    ctr = False
    for i in range (1, rows+1):
        if not ctr:     
            for j in range(1, columns+1):
                #print(str(sheet1.cell(i,j).value))
                for x in lst:
                    if x in str(sheet1.cell(i,j).value).lower():
                        
                        ctr = True
                        k = i
                        break
    if ctr:
        print(workbook.active,ii)
        pos_list.append(k)

print(pos_list)

def r_col(c, sh_num):
    workbook.active = sh_num
    sheet1 = workbook[workbook.active.title]
    rows = sheet1.max_row
    lst = []
    for i in range(1, rows + 1):
        if sheet1.cell(i, 3).value not in lst:
            lst.append(sheet1.cell(i, 3).value)
    lst1 = lst[2:]
    return lst1

def read_sheet_head_and_pos(sh_num, post):
    workbook.active = sh_num
    sheet1 = workbook[workbook.active.title]
    rows = sheet1.max_row
    pos = {}
    head = {}
    lng = {}
    col3 = r_col(3, sh_num)
    print("##", col3)
    for i in range(post+1, rows + 1):
        if sheet1.cell(i,1).value != None:
            ll = str(sheet1.cell(i,1).value).split('.')
            if len(ll) > 1 and len(ll[1]) > 0:
                pos[i] = sheet1.cell(i, 1).value
            else:
                head[i] = sheet1.cell(i, 1).value
    for k, v in pos.items():
        lng[k] = []
        for i in range(k, rows + 1):            
            if str(sheet1.cell(i, 3).value) in col3:
                lng[k].append(i)
                print(sheet1.cell(i, 3).value, i,lng[k])
                if i + 1 in pos:
                    break
    return pos, head, lng

p, h, l = read_sheet_head_and_pos(1,4)
print(p)
print(h)
print(l)
def read_sheet(sh_num, p, h):
    workbook.active = sh_num
    sheet1 = workbook[workbook.active.title]
    rows = sheet1.max_row
    it = {}
    for k,v in p.items():
       pass
    return

print(r_col(3,1))


