import openpyxl
from pyrsistent import v
from sympy import li


workbook = openpyxl.load_workbook(r'C:\Users\Rajo\Develop\Excel\primer ponuda.xlsx')
sheets = workbook.sheetnames 

#for x in sheets:
workbook.active = 2

print(sheets)
print(workbook.active.title)
sheet1 = workbook[workbook.active.title]
print(sheet1)
print(sheet1.cell(4, 3).value)

rows = sheet1.max_row
columns = sheet1.max_column

for i in range (1, rows+1):
    for j in range(1, columns+1):
        if sheet1.cell(i,j).value == 'OPIS RADOVA':
            print(i,j)

redni=[]
pozicija = []
opis = []
for i in range(5, rows+1):
    a = ''
    if sheet1.cell(i,1).value != None:
        redni.append(sheet1.cell(i,1).value)

        opis.append(sheet1.cell(i,2).value)

redni2=[]
opis2=[]
jed_mere=[]
kolicina=[]
cena=[]
iznos=[]

for i in range(6, rows+1):
    redni2.append(sheet1.cell(i,1).value)
    opis2.append(sheet1.cell(i,2).value)
    jed_mere.append(sheet1.cell(i,3).value)
    kolicina.append(sheet1.cell(i,4).value)
    cena.append(sheet1.cell(i,5).value)
    iznos.append(sheet1.cell(i,6).value)



for i in range(1,50):
    opis2.append(None)
    
print(redni2)
opis3 =[]
matrica = [ [None,None,None,None,None,None,None,None],  #redni
            [None,None,None,None,None,None,None,None],  #opis
            [None,None,None,None,None,None,None,None],  #jed
            [None,None,None,None,None,None,None,None],  #kolicina
            [None,None,None,None,None,None,None,None],  #cena
            [None,None,None,None,None,None,None,None] ] #ukcena


for i in range(len(redni2)-2):
    a = ""
    if redni2[i] != None:
        a = str(opis2[i])
        for j in range(1, 10):
            if redni2[i+j] == None:
                a = a + str(opis2[i+j])
                opis3.append(a)
            else:
                break
          
            print(a)
            
print(opis3[10])
lista=[]
for i in range(len(redni2)-2):
    matrica = [ [None,None,None,None,None,None,None,None],  #redni
            [None,None,None,None,None,None,None,None],  #opis
            [None,None,None,None,None,None,None,None],  #jed
            [None,None,None,None,None,None,None,None],  #kolicina
            [None,None,None,None,None,None,None,None],  #cena
            [None,None,None,None,None,None,None,None] ] #ukcena

    if redni2[i] != None:
        matrica[0][0] = redni2[i]
        matrica[1][0] = opis2[i]
        for j in range(1, 8):
            if redni2[i+j] == None:
                matrica[1][j] = opis2[i+j]
                matrica[2][j] = jed_mere[i+j]
                matrica[3][j] = kolicina[i+j]
                matrica[4][j] = cena[i+j]
                matrica[5][j] = iznos[i+j]

            else:
                break
        lista.append(matrica)

print(lista[15])    

for i in range(0,5):
    print(lista[15][i])    

print(lista[15][1][2])
print(lista[15][4][2])

lista2=[]

for i in range(len(lista)):
    osnOpis = str(lista[i][1][0])
    dodOpis = []
    
    for j in range(1,6):        
        if lista[i][2][j] == None:
            osnOpis = osnOpis + str(lista[i][1][j])
            
        else:
            dodOpis.append(lista[i][1][j])
        print(dodOpis)
    

   